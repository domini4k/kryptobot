# Kryptobot 1.0, pawel dominiak
#
# Skrypt osiąga zyski na podstawie delikatnych wahań kursu. Sprzedaje gdy aktualna cena jest wyższa od średniej
# dobowej, a kupuje gdy jest niższa od średniej dobowej. Dane o kupnie, sprzedazy i przebiegu kursu wykresu
# odnotowywane są w arkuszu kalkulacyjnym wraz z utworzeniem wykresu ceny.
#
# Skrypt działa na kilku zmiennych globalnych, które wykorzystywane i modyfikowane są w różnych funkcjach.
#
# Na poczatku definiujemy wartosci na ktorych skrypt bedzie dzialal: przedzialy kupna, sprzedazy i ilosc kryptowaluty
# przeznaczonej do handlu
#
# Skrypt dziala na podstawie kwoty referencji pelnego portfela. Odnosi sie do kwoty ustalonej przez uzytkowanika wobec
# ktorej ma kupowac lub sprzedawac. Np. Pelny portfel to 21 jednostek. W przypadku gdy pelny portfel jest >=21 skrypt
# sprzedaje. W przeciwnym przypadku kupuje.

import datetime, openpyxl
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl.chart import LineChart, Reference
from datetime import time
from datetime import datetime
from os import listdir
import time


dolnyProcent = 0.35
gornyProcent = 0.65
marker_kupna = False
marker_sprzedazy = False
#
browser = webdriver.Chrome()
data = datetime.now()


# otwiera Chrome, kieruje do strony giełdy
def otworz_chrome():
    import time
    browser.get('https://auth.bitbay.net/login')
    time.sleep(3)
    login = browser.find_element_by_id('email')
    login.send_keys('dominiak123@gmail.com')
    login.send_keys(Keys.ENTER)
    #
    # Uzytkownik wpisuje hasło
    #
    status = ''
    while status != 'y':
        status = input('Zalogowałeś się? Jeśli tak wpisz: y \n')
    #
    browser.get('https://app.bitbay.net/market/lsk-pln')
    time.sleep(4)


# Zebranie zmiennych takich jak najwyzsza i najnizsza wartosc dobowa, obecna wartosc. Wpisane zostaja do arkusza
# i na ich podstawie skrypt podejmuje decyzje o kupnie lub sprzedazy.
def zbierz_dane():
    #
    global lp, marker_kupna, marker_sprzedazy, ostatnia, ostatnia_cena_sprzedazy
    data = datetime.now()
    godzina = data.strftime("%X")
    ostatnia = float(browser.find_element_by_css_selector('#last > div > value').text)
    najwyzsza = float(browser.find_element_by_css_selector('#highest > value').text)
    najnizsza = float(browser.find_element_by_css_selector('#lowest > value').text)
    #
    arkusz = openpyxl.load_workbook(
        'dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
    sheet = arkusz['Sheet']
    # wpisanie danych
    while sheet['B' + str(lp)].value is not None:
        lp = lp + 1
    #
    sheet['A' + str(lp)].value = godzina
    sheet['B' + str(lp)].value = ostatnia
    sheet['C' + str(lp)].value = najwyzsza
    sheet['D' + str(lp)].value = najnizsza
    #
    # Utworzony zostaje wykres na osobnej zakladce arkusza przedstawiajacy zaleznosc wartosci od czasu
    wykres = LineChart()
    wykres.title = "Wykres ceny"
    wykres.y_axis.title = 'Cena w zł'
    wykres.x_axis.title = 'Czas'
    daneWykresu = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=lp)
    wykres.add_data(daneWykresu, titles_from_data=True)
    godziny = Reference(sheet, min_col=1, min_row=2, max_row=lp)
    wykres.set_categories(godziny)
    arkusz.remove(arkusz['wykres'])
    cs = arkusz.create_sheet('wykres')
    cs.add_chart(wykres, "B2")
    #
    arkusz.save('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
    #
    roznica = najwyzsza - najnizsza
    #
    # Warunki logiczne. Jesli odpowiednie wartosci sa spelnione funkcja sprzedaj() lub kup() ma pozwolenie na dzialanie.
    # Oprocz warunkow odpowiedzialnych za zysk istnieje jeszcze warunek pozwalajacy okreslic czy uzytkownik posiada
    # odpowiednia ilosc srodkow.
    if ostatnia >= gornyProcent * roznica + najnizsza:
        marker_sprzedazy = True
        marker_kupna = False
    elif ostatnia <= dolnyProcent * roznica + najnizsza:
        marker_sprzedazy = False
        marker_kupna = True
    else:
        marker_sprzedazy = False
        marker_kupna = False

    if ostatnia_cena_sprzedazy > 0 and ostatnia_cena_sprzedazy >= ostatnia and ostatnia <= dolnyProcent * roznica + najnizsza:
        marker_sprzedazy = False
        marker_kupna = True
        ostatnia_cena_sprzedazy = 0

    if float(browser.find_element_by_css_selector(
            '#wallet-top-h-trigger > div > div > second-value').text[0:len(browser.find_element_by_css_selector(
            '#wallet-top-h-trigger > div > div > second-value').text)-4]) < ostatnia * 2:
        marker_kupna = False


def sprzedaj():
    while True:
        zbierz_dane()
        if marker_sprzedazy:
            # Zlozenie oferty. Jesli skrypt uzna ze jest to obecny czas na sprzedaz, sklada oferte i czeka az sprzedaz
            # faktycznie nastapi. Nastepnie zapisuje wynik do arkusza.
            ilosc_lsk = browser.find_element_by_css_selector('#amount-sell')
            ilosc_lsk.clear()
            ilosc_lsk.send_keys(2)
            cena = browser.find_element_by_css_selector('#price-sell')
            cena.clear()
            cena.send_keys(str(ostatnia - 0.02))
            przycisk_sprzedaj = browser.find_element_by_css_selector('#sell-button')
            time.sleep(1)
            while True:
                try:
                    przycisk_sprzedaj.click()
                    break
                except:
                    time.sleep(1)
            time.sleep(10)

            # skrypt czeka aż oferta zostanie wypełniona
            while True:
                if not browser.find_elements_by_css_selector(
                        '#market-my-offers > div.content-wrapper > div.list.with-types > div > div'):
                    print('Sprzedano po cenie: ' + str(ostatnia - 0.02))
                    break
                else:
                    time.sleep(15)

            # zapisanie sprzedazy do arkusza
            arkusz = openpyxl.load_workbook(
                'dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
            sheet = arkusz['Sheet']
            sheet['E' + str(lp)].value = ostatnia - 0.02
            sheet['F' + str(lp)].value = 'Sprzedano'
            arkusz.save('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
            # Po spelnieniu wszystkich warunkow wychodzi z petli.
            break
        else:
            time.sleep(15)


# Funkcja analogiczna do sprzedaj()
def kup():
    while True:
        zbierz_dane()
        # zlozenie oferty
        if marker_kupna:
            ilosc_lsk = browser.find_element_by_css_selector('#amount')
            ilosc_lsk.clear()
            ilosc_lsk.send_keys("2.00863714")
            cena = browser.find_element_by_css_selector('#price')
            cena.clear()
            cena.send_keys(str(ostatnia))
            przycisk_kup = browser.find_element_by_css_selector('#buy-button')
            time.sleep(5)
            while True:
                try:
                    przycisk_kup.click()
                    break
                except:
                    time.sleep(1)
            time.sleep(10)
            #
            while True:
                if not browser.find_elements_by_css_selector(
                        '#market-my-offers > div.content-wrapper > div.list.with-types > div > div'):
                    print('Kupiono po cenie: ' + str(ostatnia))
                    break
                else:
                    time.sleep(15)
            #
            arkusz = openpyxl.load_workbook(
                'dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
            sheet = arkusz['Sheet']
            sheet['E' + str(lp)].value = ostatnia
            sheet['F' + str(lp)].value = 'Kupiono'
            arkusz.save('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
            break
        else:
            time.sleep(15)


#
# Start #
#
# Utworzenie pliku arkusza w folderze 'dane'
if not Path(
        'C:/Users/Dominiak/PycharmProjects/kryptobot/dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(
                data.year) + '.xlsx').is_file():
    # utworzenie arkusza w przypadku braku
    data = datetime.now()
    godzina = data.strftime("%X")
    arkusz = openpyxl.Workbook('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
    arkusz.save('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
    arkusz = openpyxl.load_workbook(
        'dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
    sheet = arkusz['Sheet']
    sheet['A1'] = 'Godzina'
    sheet['B1'] = 'Ostatnia wartość'
    sheet['C1'] = 'Najwyzsza wartość dobowa'
    sheet['D1'] = 'Najnizsza wartość dobowa'
    cs = arkusz.create_sheet('wykres')
    #
    arkusz.save('dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
#
# Otwiera strone gieldy, uzytkowanik loguje sie
otworz_chrome()
#
# W przypadku gdy sprzedaz nastapila poprzedniego dnia i uzytkownik nie posiada pelnego portfela skrypt musi dowiedziec
# sie jaka byla cena sprzedazy, aby nie spowodowac strat obecnym kursem. W tym wypadku przeszukuje arkusz z obecnego
# dnia oraz jednej daty poprzedniej. W przypadku dawniejszych operacji nalezy kwote wpisac recznie do terminala aby
# uniknac strat.
ostatnia_cena_sprzedazy = 0
lp = 2
if float(browser.find_element_by_css_selector('#wallet-top-h-trigger > div > div > first-value').text[0:9]) < 20:

    if Path('C:/Users/Dominiak/PycharmProjects/kryptobot/dane/arkusz ' + str(data.day) + '.' + str(data.month) +
            '.' + str(data.year) + '.xlsx').is_file():
        arkusz = openpyxl.load_workbook(
            'dane/arkusz ' + str(data.day) + '.' + str(data.month) + '.' + str(data.year) + '.xlsx')
        sheet = arkusz['Sheet']
        while sheet['B' + str(lp)].value is not None:
            lp = lp + 1
        while sheet['F' + str(lp)].value != 'Sprzedano':
            lp = lp - 1
            if lp == 1:
                sciezka = listdir('C:/Users/Dominiak/PycharmProjects/kryptobot/dane')
                sciezka.sort(reverse=True)
                arkusz = openpyxl.load_workbook('dane/' + str(sciezka[1]))
                sheet = arkusz['Sheet']
                while sheet['B' + str(lp)].value is not None:
                    lp = lp + 1
                while sheet['F' + str(lp)].value != 'Sprzedano':
                    lp = lp - 1
                    if lp == 1:
                        status = 0
                        while status == 0:
                            status = input('Wpisz ostatnia cene sprzedazy \n')
                        ostatnia_cena_sprzedazy = status
        if ostatnia_cena_sprzedazy == 0:
            ostatnia_cena_sprzedazy = sheet['E' + str(lp)].value
        lp = 2
        print('Ostatnia cena sprzedazy to: ' + str(ostatnia_cena_sprzedazy))

# Skrypt wchodzi w nieskonczona petle wlasciwa. Wymienne zachodza operacje kupna i sprzedazy jesli tylko istenieja
# potrzebne warunki.
while True:
    #
    status_portfela = float(
        browser.find_element_by_css_selector('#wallet-top-h-trigger > div > div > first-value').text[0:9])
    #
    if status_portfela > 20:
        sprzedaj()
    else:
        kup()
